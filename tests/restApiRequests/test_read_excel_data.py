import os
import requests
import json
import openpyxl

current_directory = os.getcwd()
root_folder_name = os.path.basename(current_directory)
parent_directory = current_directory.rstrip(root_folder_name)

def test_add_student_details_using_excel():
    API_URLS = "https://thetestingworldapi.com/api/studentsDetails"
    file = open(parent_directory + "data/studentDetailsExcel.json", 'r')
    json_request = json.loads(file.read())
    wb = openpyxl.load_workbook(parent_directory+"data/testData.xlsx")
    sheet = wb['Sheet1']
    rows = sheet.max_row
    for i in range(1,rows+1):
        cell_firstName = sheet.cell(row=i,column=1)
        cell_midName = sheet.cell(row=i, column=2)
        cell_LastName = sheet.cell(row=i, column=3)
        cell_dob = sheet.cell(row=i, column=4)
        json_request['first_name'] = cell_firstName.value
        json_request['middle_name'] = cell_midName.value
        json_request['last_name'] = cell_LastName.value
        json_request['date_of_birth'] = cell_dob.value
        response = requests.post(API_URLS, json_request)
        print(response.text)
        assert  response.status_code == 201

