import openpyxl

class Common:
    def __init__(self, filepath, sheetName):
        self.wb = openpyxl.load_workbook(filepath)
        self.sheet = self.wb[sheetName]

    def fetch_row_count(self):
        rows = self.sheet.max_row
        return rows

    def fetch_column_count(self):
        col = self.sheet.max_column
        return col

    def fetch_key_names(self):
        col = self.sheet.max_column
        li = []
        for i in range(1, col + 1):
            cell = self.sheet.cell(row=1, column=i)
            li.insert(i - 1, cell.value)

        return li

    def update_request_with_data(self, rowNumber, json_request, keylist):
        col = self.sheet.max_column
        for i in range(1, col + 1):
            cell = self.sheet.cell(row=rowNumber, column=i)
            json_request[keylist[i - 1]] = cell.value

        return json_request
